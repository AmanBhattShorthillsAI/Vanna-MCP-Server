import os
import json
import weaviate
import pandas as pd
from vanna.weaviate.weaviate_vector import WeaviateDatabase
from vanna.base import VannaBase
from dotenv import load_dotenv
from langchain_openai import AzureChatOpenAI
from langchain_core.messages import SystemMessage, HumanMessage, AIMessage
from contextlib import asynccontextmanager
from collections.abc import AsyncIterator
from dataclasses import dataclass
import time
import openpyxl
from openpyxl import load_workbook

# Import MCP and anyio components
from mcp.server.fastmcp import FastMCP, Context
import anyio

# Import and configure the logging module
import logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

import signal
import sys

def shutdown_handler(signum, frame):
    logging.info("Received shutdown signal. Cleaning up MCP server...")
    try:
        sys.exit(0)
    except Exception as e:
        logging.error(f"Error during shutdown: {e}")
        sys.exit(1)

# Register signal handlers
signal.signal(signal.SIGINT, shutdown_handler)
signal.signal(signal.SIGTERM, shutdown_handler)

# Load environment variables from .env file
load_dotenv()

class LangChainAzureChat(VannaBase):
    def __init__(self, config=None):
        super().__init__(config=config)
        self.llm = AzureChatOpenAI(
            azure_deployment="gpt-4.1",
            api_version="2024-02-15-preview",
            temperature=0.0,
            max_tokens=1000,
            api_key=os.getenv("OPENAI_API_KEY"),
            azure_endpoint=os.getenv("AZURE_OPENAI_ENDPOINT"),
        )

    def system_message(self, message: str) -> SystemMessage:
        return SystemMessage(content=message)

    def user_message(self, message: str) -> HumanMessage:
        return HumanMessage(content=message)

    def assistant_message(self, message: str) -> AIMessage:
        return AIMessage(content=message)

    def submit_prompt(self, prompt, **kwargs) -> str:
        response = self.llm.invoke(prompt)
        logging.info(f"Response: {response}")
        input_tokens = response.usage_metadata.get('input_tokens')
        output_tokens = response.usage_metadata.get('output_tokens')
        return response.content, input_tokens, output_tokens
        
class MyVanna(WeaviateDatabase, LangChainAzureChat):
    def __init__(self, config=None):
        self.config = config or {}
        WeaviateDatabase.__init__(self, config=config)
        LangChainAzureChat.__init__(self, config=config)

    def _initialize_weaviate_client(self):
        if self.config.get("weaviate_api_key"):
            return weaviate.connect_to_weaviate_cloud(
                cluster_url=self.config["weaviate_url"],
                auth_credentials=weaviate.auth.AuthApiKey(self.config["weaviate_api_key"]),
                skip_init_checks=True
            )
        else:
            raise ValueError("Weaviate API key is required for online Weaviate.")

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.weaviate_client.close()
        logging.info("\nWeaviate connection closed successfully.")

# --- MCP Server Integration ---

@dataclass
class AppContext:
    vn: MyVanna

@asynccontextmanager
async def app_lifespan(server: FastMCP) -> AsyncIterator[AppContext]:
    """
    Manages the Vanna AI instance's lifecycle.
    """
    logging.info("🚀 MCP Server starting up...")
    config = {
        "weaviate_url": os.getenv("WEAVIATE_URL"),
        "weaviate_api_key": os.getenv("WEAVIATE_API_KEY"),
    }
    
    with MyVanna(config=config) as vn:
        logging.info("🔗 Connecting to SQLite database...")
        vn.connect_to_sqlite("financial.sqlite")
        # The server should not train on startup.
        # Run the `train.py` script once to populate your vector store.
        logging.info("✅ Vanna AI is ready. Server is online.")
        yield AppContext(vn=vn)
        
    logging.info("🔌 MCP Server shutting down...")

mcp = FastMCP(
    name="VannaAIServer",
    instructions="A server that uses Vanna AI to answer questions about a financial database.",
    lifespan=app_lifespan
)

# Excel logging utility
LOG_FILE = "query_log.xlsx"
LOG_COLUMNS = [
    "question", "prompt", "llm_input_tokens", "llm_output_tokens", "llm_cost", "sql_gen_time", "sql_query", "fetch_time", "fetch_result"
]

def append_log_to_excel(row_dict):
    if not os.path.exists(LOG_FILE):
        # Create new DataFrame and save
        df = pd.DataFrame([row_dict], columns=LOG_COLUMNS)
        df.to_excel(LOG_FILE, index=False)
    else:
        # Append to existing file
        wb = load_workbook(LOG_FILE)
        ws = wb.active
        ws.append([row_dict.get(col, "") for col in LOG_COLUMNS])
        wb.save(LOG_FILE)

def calculate_cost(input_tokens, output_tokens):
    # Example cost calculation (adjust rates as needed)
    if input_tokens is None or output_tokens is None:
        return None
    return (input_tokens * 0.0000015 + output_tokens * 0.000006) * 0.000001

# **FIX: This is the corrected ask_sql tool**
@mcp.tool()
async def ask_sql(question: str, ctx: Context) -> str:
    """
    Takes a natural language question about financial data and returns a SQL query.
    Logs question, prompt, LLM tokens, cost, and timing to Excel.
    """
    try:
        vn_instance = ctx.request_context.lifespan_context.vn
        logging.info(f"Received question for SQL generation: '{question}'")
        log_row = {"question": question}

        def generate_sql_with_full_context(vn: MyVanna, q: str):
            question_sql_list = vn.get_similar_question_sql(q)
            prompt = vn.get_sql_prompt(
                initial_prompt=vn.config.get("initial_prompt", None) if hasattr(vn, "config") else None,
                question=q,
                question_sql_list=question_sql_list,
                ddl_list=vn.get_related_ddl(q),
                doc_list=vn.get_related_documentation(q)
            )
            logging.info(f"Prompt: {prompt}")
            llm_start = time.time()
            sql, input_tokens, output_tokens = vn.submit_prompt(prompt)
            llm_cost = calculate_cost(input_tokens, output_tokens)
            llm_time = time.time() - llm_start
            return sql, prompt, input_tokens, output_tokens, llm_cost, llm_time

        sql_query, prompt, input_tokens, output_tokens, llm_cost, llm_time = await anyio.to_thread.run_sync(
            generate_sql_with_full_context, vn_instance, question
        )
        logging.info(f"Generated SQL: {sql_query}")
        prompt_str = "\n".join([f"({msg.type}) {msg.content}" for msg in prompt])
        log_row.update({
            "prompt": prompt_str,
            "llm_input_tokens": input_tokens,
            "llm_output_tokens": output_tokens,
            "llm_cost": llm_cost,
            "sql_gen_time": llm_time,
            "sql_query": sql_query
        })
        append_log_to_excel(log_row)
        return sql_query or "Could not generate a valid SQL query."
    except Exception as e:
        logging.error(f"Error in ask_sql tool: {e}", exc_info=True)
        return f"Error generating SQL query: {e}"


@mcp.tool()
async def run_sql(sql_query: str, ctx: Context) -> str:
    """
    Executes a SQL query against the financial database and returns the result as a JSON string.
    Logs fetch time and result to Excel.
    """
    try:
        vn_instance = ctx.request_context.lifespan_context.vn
        logging.info(f"Executing SQL query: {sql_query}")
        fetch_start = time.time()
        df = await anyio.to_thread.run_sync(vn_instance.run_sql, sql_query)
        fetch_time = time.time() - fetch_start
        if df is not None:
            fetch_result = df.to_json(orient='records')
        else:
            fetch_result = "Query executed, but no results were returned."
        # Try to append to the last row in Excel (if possible)
        try:
            if os.path.exists(LOG_FILE):
                wb = load_workbook(LOG_FILE)
                ws = wb.active
                last_row = ws.max_row
                ws.cell(row=last_row, column=LOG_COLUMNS.index("fetch_time")+1, value=fetch_time)
                ws.cell(row=last_row, column=LOG_COLUMNS.index("fetch_result")+1, value=fetch_result)
                wb.save(LOG_FILE)
        except Exception as e:
            logging.error(f"Error updating fetch log in Excel: {e}")
        return fetch_result
    except Exception as e:
        logging.error(f"Error in run_sql tool: {e}", exc_info=True)
        return f"Error executing SQL query: {e}"

if __name__ == '__main__':
    logging.info("Starting Vanna AI MCP Server...")
    mcp.run()